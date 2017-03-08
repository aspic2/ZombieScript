#MT: Handles reading and writing Excel spreadsheet info

from openpyxl import load_workbook
from openpyxl import Workbook
from os import getcwd

'''To make a new workbook, set the second parameter 'assemble=' to True
To use an existing notebook, provide a path
'''
class Spreadsheet(object):
    #name will be used for the saved file
    def __init__(self, name='result', assemble=False, path=None):
        self.name = name
        self.assemble = assemble
        self.path = path
        if self.assemble:
            self.wb = Workbook(write_only=True)
            self.ws = self.wb.create_sheet()
            #Column names. Should correspond with values of token in main()
            title = ("LID", 'Name', 'Status', 'isMissingCreatives')
            self.ws.append(title)
            self.assemble = False

    def read(self, column):
        wb = load_workbook(self.path)
        ws = wb.active
        self.values = []
        for v in ws[column]:
            try:
                int(v.value)
                self.values.append(v.value)
            except Exception as ValueError:
                continue
        return self.values

    def append(self, values):
        self.ws.append(values)

    def save(self):
        file_name = str(getcwd() + '\\outputfiles\\' + self.name + '.xlsx')
        self.wb.save(filename=file_name)
